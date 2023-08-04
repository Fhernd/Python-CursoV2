import datetime
import json
import pickle
import os
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from tkinter import filedialog
import threading

from tkcalendar import Calendar, DateEntry

import openpyxl

import flet as ft
from flet import AppBar, ElevatedButton, FilledButton, Page, PopupMenuButton, PopupMenuItem, Text, View, colors

from modelos.inventario import Inventario
from modelos.producto import Producto
from modelos.venta import Venta
from conexion import conectar


class VentanaPrincipal(tk.Frame):
    def __init__(self, parent=None):
        super().__init__()
        self.parent = parent

        self.inicializar_gui()
        self.cargar_inventario()

    def inicializar_gui(self):
        self.parent.title("Gestor Inventario - Dispositivos S.a.s.")

        self.menu_bar = tk.Menu(self)
        
        self.file_menu = tk.Menu(self.menu_bar, tearoff=0)
        self.file_menu.add_command(label="Salir", command=self.quit)
        self.file_menu.add_command(label="Exportar datos a CSV...", command=self.exportar_csv)
        self.file_menu.add_command(label="Exportar datos a JSON...", command=self.exportar_datos_json)
        self.file_menu.add_command(label="Exportar datos a XLSX...", command=self.exportar_datos_xlsx)
        self.menu_bar.add_cascade(label="Archivo", menu=self.file_menu)

        self.product_menu = tk.Menu(self.menu_bar, tearoff=0)
        self.product_menu.add_command(label="Registrar", command=self.registrar_producto)
        self.product_menu.add_command(label="Listar", command=self.listar_productos)
        self.product_menu.add_command(label="Vender", command=self.vender_producto)
        self.product_menu.add_command(label="Buscar", command=self.buscar_producto)
        self.product_menu.add_command(label="Cambiar disponibilidad", command=self.cambiar_disponibilidad_producto)
        self.menu_bar.add_cascade(label="Productos", menu=self.product_menu)
        self.report_menu = tk.Menu(self.menu_bar, tearoff=0)
        self.report_menu.add_command(label="Ventas en un rango de fechas", command=self.reporte_ventas_rango_fechas)
        self.report_menu.add_command(label="Top 5 de productos más vendidos", command=self.reporte_top_5_productos_mas_vendidos)
        self.report_menu.add_command(label="Top 5 de productos menos vendidos", command=self.reporte_top_5_productos_menos_vendidos)
        self.menu_bar.add_cascade(label="Reportes", menu=self.report_menu)
        self.help_menu = tk.Menu(self.menu_bar, tearoff=0)
        self.help_menu.add_command(label="Acerca de", command=self.acerca_de)
        self.menu_bar.add_cascade(label="Ayuda", menu=self.help_menu)
        self.parent.config(menu=self.menu_bar)

        self.parent.protocol("WM_DELETE_WINDOW", self.salir)
    
    def salir(self):
        if messagebox.askokcancel("Salir", "¿Desea salir de la aplicación?"):
            self.inventario.guardar_datos()
            self.parent.destroy()

    def exportar_csv(self):
        archivo = filedialog.asksaveasfile(title="Exportar productos a CSV", defaultextension=".csv", filetypes=(("CSV", "*.csv"),))

        if archivo is not None:
            with open(archivo.name, 'w', newline='', encoding='utf8') as f:
                f.write('Código,Nombre,Precio,Cantidad,Disponible\n')

                for producto in self.inventario.productos:
                    f.write(f'{producto.codigo},{producto.nombre},{producto.precio},{producto.cantidad},{"Sí" if producto.disponible else "No"}\n')
                
                messagebox.showinfo("Exportar a CSV", "Datos de productos exportados correctamente.")
        
        archivo = filedialog.asksaveasfile(title="Exportar ventas a CSV", defaultextension=".csv", filetypes=(("CSV", "*.csv"),))

        if archivo is not None:
            with open(archivo.name, 'w', newline='', encoding='utf8') as f:
                f.write('Código Producto,Cantidad,Total sin IVA,Fecha\n')

                for venta in self.inventario.ventas:
                    f.write(f'{venta.codigo_producto},{venta.cantidad},{venta.total_sin_iva},{venta.fecha.strftime("%d/%m/%Y")}\n')
                
                messagebox.showinfo("Exportar a CSV", "Datos de ventas exportados correctamente.")
        
    def exportar_datos_json(self):
        archivo = filedialog.asksaveasfile(title="Exportar datos a JSON", defaultextension=".json", filetypes=(("JSON", "*.json"),))

        if archivo is not None:
            datos = {
                'productos': [e.__dict__ for e in self.inventario.productos],
                'ventas': [
                    {
                        'codigo_producto': e.codigo_producto,
                        'cantidad': e.cantidad,
                        'total_sin_iva': e.total_sin_iva,
                        'fecha': e.fecha.strftime("%d/%m/%Y")
                    } for e in self.inventario.ventas
                ]
            }

            with open(archivo.name, 'wt', encoding='utf8') as f:
                # Serializar a JSON todos los productos y ventas:
                json.dump(datos, f, indent=4)
            
            messagebox.showinfo("Exportar a JSON", "Datos exportados correctamente.")
    
    def exportar_datos_xlsx(self):
        # Solicitar ruta de archivo:
        archivo = filedialog.asksaveasfile(title="Exportar datos a XLSX", defaultextension=".xlsx", filetypes=(("XLSX", "*.xlsx"),))

        if archivo is not None:
            # Crear libro de trabajo:
            libro_trabajo = openpyxl.Workbook()

            # Crear hoja de productos:
            hoja_productos = libro_trabajo.create_sheet("Productos")
            hoja_productos.append(['Código', 'Nombre', 'Precio', 'Cantidad', 'Disponible'])

            # Agregar productos a la hoja:
            for producto in self.inventario.productos:
                hoja_productos.append([producto.codigo, producto.nombre, producto.precio, producto.cantidad, "Sí" if producto.disponible else "No"])
            
            # Crear hoja de ventas:
            hoja_ventas = libro_trabajo.create_sheet("Ventas")
            hoja_ventas.append(['Código Producto', 'Cantidad', 'Total sin IVA', 'Fecha'])

            # Agregar ventas a la hoja:
            for venta in self.inventario.ventas:
                hoja_ventas.append([venta.codigo_producto, venta.cantidad, venta.total_sin_iva, venta.fecha.strftime("%d/%m/%Y")])
            
            # Guardar libro de trabajo:
            libro_trabajo.save(archivo.name)

            messagebox.showinfo("Exportar a XLSX", "Datos exportados correctamente.")

    def registrar_producto(self):
        registro_producto_frame = ProductoCrearFrame(self.parent, self.inventario)
        registro_producto_frame.grab_set()

    def listar_productos(self):
        listar_productos_frame = ProductosListarFrame(self.parent, self.inventario)
        listar_productos_frame.grab_set()

    def vender_producto(self):
        venta_producto_frame = ProductoVenderFrame(self.parent, self.inventario)
        venta_producto_frame.grab_set()

    def buscar_producto(self):
        buscar_producto_frame = ProductoBuscarFrame(self.parent, self.inventario)
        buscar_producto_frame.grab_set()
    
    def cambiar_disponibilidad_producto(self):
        cambiar_disponibilidad_producto_frame = ProductoCambiarDisponibilidadFrame(self.parent, self.inventario)
        cambiar_disponibilidad_producto_frame.grab_set()

    def reporte_ventas_rango_fechas(self):
        reporte_ventas_rango_fechas_frame = ReporteVentasRangoFechasFrame(self.parent, self.inventario)
        reporte_ventas_rango_fechas_frame.grab_set()

    def reporte_top_5_productos_mas_vendidos(self):
        ventana = Top5VendidosFrame(self.parent, self.inventario, self.inventario.top_5_mas_vendidos())
        ventana.grab_set()

    def reporte_top_5_productos_menos_vendidos(self):
        ventana = Top5VendidosFrame(self.parent, self.inventario, self.inventario.top_5_menos_vendidos())
        ventana.grab_set()

    def cargar_inventario(self):
        """
        Carga el inventario desde el archivo inventario.pickle si es que existe.
        """
        if os.path.isfile('inventario/inventario.pickle'):
            self.inventario = Inventario()
            
            if messagebox.askyesno('Mensaje', '¿Desea cargar el inventario?'):
                with open('inventario/inventario.pickle', 'rb') as f:
                    resultado = pickle.load(f)
                
                if resultado:
                    self.inventario.productos = self.cargar_productos(resultado['productos']) if 'productos' in resultado else []
                    self.inventario.ventas = self.cargar_ventas(resultado['ventas']) if 'ventas' in resultado else []

    def cargar_productos(self, productos):
        """
        Carga los productos en el inventario.

        Parameters:
        productos: lista de productos a cargar en el inventario.

        Returns:
        Lista de productos cargados en el inventario.
        """
        productos_inventario = []

        for p in productos:
            producto = Producto(p.codigo, p.nombre, p.precio, p.cantidad, p.disponible)
            productos_inventario.append(producto)
        
        return productos_inventario

    def cargar_ventas(self, ventas):
        """
        Carga las ventas en el inventario.

        Parameters:
        ventas: lista de ventas a cargar en el inventario.

        Returns:
        Lista de ventas cargadas en el inventario.
        """
        ventas_inventario = []

        for v in ventas:
            venta = Venta(v.codigo_producto, v.cantidad, v.total_sin_iva, v.fecha)
            ventas_inventario.append(venta)
        
        return ventas_inventario

    def acerca_de(self):
        # Incluir un mensaje con el desarrollador del software, la versión y el sitio Web:
        mensaje = 'Desarrollado por: John Ortiz Ordoñez\n' + \
            'Versión: 1.0.0\n' + \
            'Sitio Web: https://ortizol.blogspot.com'
        
        messagebox.showinfo('Acerca de', mensaje)

class ProductoCrearFrame(tk.Toplevel):
    def __init__(self, parent, inventario):
        tk.Toplevel.__init__(self, parent)

        self.inventario = inventario

        self.inicializar_gui()

    def inicializar_gui(self):
        self.title('Producto - Crear')
        
        # Variable para almacenar el inventario:
        self.codigo_var = tk.IntVar()
        codigo_label = tk.Label(self, text='Código:')
        codigo_label.grid(row=0, column=0, padx=5, pady=5, sticky=tk.W)
        codigo_entry = tk.Entry(self, textvariable=self.codigo_var)
        codigo_entry.grid(row=0, column=1, padx=5, pady=5)
        
        self.nombre_var = tk.StringVar()
        nombre_label = tk.Label(self, text='Nombre:')
        nombre_label.grid(row=1, column=0, padx=5, pady=5, sticky=tk.W)
        nombre_entry = tk.Entry(self, textvariable=self.nombre_var)
        nombre_entry.grid(row=1, column=1, padx=5, pady=5)

        self.precio_var = tk.DoubleVar()
        precio_label = tk.Label(self, text='Precio:')
        precio_label.grid(row=2, column=0, padx=5, pady=5, sticky=tk.W)
        precio_entry = tk.Entry(self, textvariable=self.precio_var)
        precio_entry.grid(row=2, column=1, padx=5, pady=5)

        self.cantidad_var = tk.IntVar()
        cantidad_label = tk.Label(self, text='Cantidad:')
        cantidad_label.grid(row=3, column=0, padx=5, pady=5, sticky=tk.W)
        cantidad_entry = tk.Entry(self, textvariable=self.cantidad_var)
        cantidad_entry.grid(row=3, column=1, padx=5, pady=5)

        self.disponible_var = tk.BooleanVar()
        disponible_checkbutton = tk.Checkbutton(self, text='Disponible para venta?', variable=self.disponible_var)
        disponible_checkbutton.grid(row=4, column=0, padx=5, pady=5, sticky=tk.W)

        # Crear el botón de crear y la función asociada
        crear_button = tk.Button(self, text='Crear', command=lambda: self.crear_producto())
        crear_button.grid(row=5, column=1, padx=5, pady=5, sticky=tk.E)

        # Establecer tamaño mínimo del formulario
        self.minsize(250, 250)

        # Hacer que el formulario se adapte al tamaño de la ventana anidada
        self.grid_rowconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        self.grid_rowconfigure(2, weight=1)
        self.grid_rowconfigure(3, weight=1)
        self.grid_rowconfigure(4, weight=1)
        self.grid_rowconfigure(5, weight=1)
        self.grid_columnconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)

    def crear_producto(self):
        codigo = self.codigo_var.get()
        nombre = self.nombre_var.get()
        precio = self.precio_var.get()
        cantidad = self.cantidad_var.get()
        disponible = self.disponible_var.get()
        
        if not codigo or not nombre or not precio or not cantidad:
            messagebox.showwarning('Mensaje', 'Todos los campos son obligatorios. Los valores numéricos deben ser mayores a 0.')
            return
        
        try:
            codigo = int(codigo)
        except ValueError:
            messagebox.showwarning('Mensaje', 'El código debe ser numérico.')
            return

        if codigo <= 0:
            messagebox.showwarning('Mensaje', 'El código debe ser positivo.')
            return
        
        try:
            precio = float(precio)
        except ValueError:
            messagebox.showwarning('Mensaje', 'El precio debe ser numérico.')
            return

        if precio <= 0:
            messagebox.showwarning('Mensaje', 'El precio debe ser positivo.')
            return
        
        try:
            cantidad = int(cantidad)
        except ValueError:
            messagebox.showwarning('Mensaje', 'La cantidad debe ser numérica.')
            return

        if cantidad <= 0:
            messagebox.showwarning('Mensaje', 'La cantidad debe ser positiva.')
            return
        
        producto = self.inventario.buscar_producto(codigo)

        if producto:
            messagebox.showwarning('Mensaje', 'Ya existe un producto con el código especificado.')
            return

        nuevo_producto = Producto(codigo, nombre, precio, cantidad, disponible)
        self.inventario.registrar_producto(nuevo_producto)

        self.inventario.guardar_datos()

        messagebox.showinfo('Mensaje', 'El producto se ha creado de forma satisfactoria.')

        self.codigo_var.set('')
        self.nombre_var.set('')
        self.precio_var.set('')
        self.cantidad_var.set('')
        self.disponible_var.set(False)
        

    def quit(self) -> None:
        """
        Cierra la aplicación.
        """
        if messagebox.askokcancel("Salir", "¿Desea salir de la aplicación?"):
            exit()

class ProductosListarFrame(tk.Toplevel):
    def __init__(self, master, inventario):
        super().__init__(master)
        self.inventario = inventario

        self.inicializar_gui()
    
    def inicializar_gui(self):
        self.title("Productos")
        
        self.table = ttk.Treeview(self)
        self.table["columns"] = ("Código", "Nombre", "Precio", "Cantidad", "Disponible")
        
        self.table.heading("Código", text="Código")
        self.table.heading("Nombre", text="Nombre")
        self.table.heading("Precio", text="Precio")
        self.table.heading("Cantidad", text="Cantidad")
        self.table.heading("Disponible", text="¿Disponible?")
        
        for producto in self.inventario.productos:
            codigo = producto.codigo
            nombre = producto.nombre
            precio = producto.precio
            cantidad = producto.cantidad
            disponible = 'Sí' if producto.disponible else 'No'
            
            self.table.insert("", tk.END, values=(codigo, nombre, precio, cantidad, disponible))
        
        # Ajustar el tamaño de las columnas
        for column in ("Código", "Nombre", "Precio", "Cantidad", "Disponible"):
            self.table.column(column, width=100, anchor=tk.CENTER)
        
        # Colocar la tabla en el layout
        self.table.pack(fill=tk.BOTH, expand=True)

class ProductoVenderFrame(tk.Toplevel):
    def __init__(self, parent, inventario):
        super().__init__(parent)
        self.title("Venta de Producto")
        self.parent = parent
        self.inventario = inventario
        
        self.inicializar_gui()
    
    def inicializar_gui(self):
        label_codigo = tk.Label(self, text="Código:")
        label_codigo.grid(row=0, column=0)
        
        self.codigo_var = tk.IntVar()
        entry_codigo = tk.Entry(self, textvariable=self.codigo_var)
        entry_codigo.grid(row=0, column=1)
        
        label_cantidad = tk.Label(self, text="Cantidad:")
        label_cantidad.grid(row=1, column=0)
        
        self.cantidad_var = tk.IntVar()
        entry_cantidad = tk.Entry(self, textvariable=self.cantidad_var)
        entry_cantidad.grid(row=1, column=1)
        
        button_vender = tk.Button(self, text="Vender", command=self.vender_producto)
        button_vender.grid(row=2, columnspan=2)
    
    def vender_producto(self):
        codigo = self.codigo_var.get()
        cantidad = self.cantidad_var.get()

        if not codigo or not cantidad:
            messagebox.showwarning("Mensaje", "Todos los campos son obligatorios. Los valores numéricos deben ser mayores a 0.")
            return

        try:
            codigo = int(codigo)
        except ValueError:
            messagebox.showwarning("Mensaje", "El código debe ser numérico.")
            return
        
        if codigo <= 0:
            messagebox.showwarning("Mensaje", "El código debe ser positivo.")
            return
        
        producto = self.inventario.buscar_producto(codigo)

        if not producto:
            messagebox.showwarning("Mensaje", "El producto no existe.")
            return

        try:
            cantidad = int(cantidad)
        except ValueError:
            messagebox.showwarning("Mensaje", "La cantidad debe ser numérica.")
            return

        if cantidad <= 0:
            messagebox.showwarning("Mensaje", "La cantidad debe ser positiva.")
            return
        
        if cantidad > producto.cantidad:
            messagebox.showwarning("Mensaje", "La cantidad solicitada es mayor a la cantidad disponible.")
            return
        
        total = producto.precio * cantidad

        venta = Venta(codigo, cantidad, total)
        self.inventario.realizar_venta(venta)
        total *= 1.19
        total = "${:,.2f}".format(total)

        messagebox.showinfo("Mensaje", f"La venta se ha realizado de forma satisfactoria. Total a pagar: {total}")

        self.inventario.guardar_datos()

        self.codigo_var.set('')
        self.cantidad_var.set('')

class ProductoBuscarFrame(tk.Toplevel):
    def __init__(self, master, inventario):
        super().__init__(master)
        self.inventario = inventario

        self.title("Producto - Buscar")

        self.inicializar_gui()

    def inicializar_gui(self):
        label_codigo = ttk.Label(self, text="Código:")
        label_nombre = ttk.Label(self, text="Nombre:")
        label_precio = ttk.Label(self, text="Precio:")
        label_cantidad = ttk.Label(self, text="Cantidad:")

        self.codigo_var = tk.IntVar()
        self.entry_codigo = ttk.Entry(self, textvariable=self.codigo_var)
        self.entry_codigo.focus()

        self.nombre_var = tk.StringVar()
        self.entry_nombre = ttk.Entry(self, textvariable=self.nombre_var)
        self.entry_nombre.state(['disabled'])

        self.precio_var = tk.DoubleVar()
        self.entry_precio = ttk.Entry(self, textvariable=self.precio_var)
        self.entry_precio.state(['disabled'])

        self.cantidad_var = tk.IntVar()
        self.entry_cantidad = ttk.Entry(self, textvariable=self.cantidad_var)
        self.entry_cantidad.state(['disabled'])

        self.disponible_var = tk.BooleanVar()
        self.checkbox_disponible = ttk.Checkbutton(self, text="Disponible para venta", variable=self.disponible_var)
        self.checkbox_disponible.state(['disabled'])

        button_buscar = ttk.Button(self, text="Buscar", command=self.buscar_producto)

        label_codigo.grid(row=0, column=0, sticky=tk.W)
        self.entry_codigo.grid(row=0, column=1)
        button_buscar.grid(row=0, column=2)

        label_nombre.grid(row=1, column=0, sticky=tk.W)
        self.entry_nombre.grid(row=1, column=1)

        label_precio.grid(row=2, column=0, sticky=tk.W)
        self.entry_precio.grid(row=2, column=1)

        label_cantidad.grid(row=3, column=0, sticky=tk.W)
        self.entry_cantidad.grid(row=3, column=1)

        self.checkbox_disponible.grid(row=4, columnspan=2, sticky=tk.W)

    def buscar_producto(self):
        codigo = self.entry_codigo.get()

        if not codigo and codigo == 0:
            messagebox.showwarning("Mensaje", "El código es obligatorio.")
            return
        
        try:
            codigo = int(codigo)
        except ValueError:
            messagebox.showwarning("Mensaje", "El código debe ser numérico.")
            return

        producto = self.inventario.buscar_producto(codigo)
        
        self.nombre_var.set('')
        self.precio_var.set('')
        self.cantidad_var.set('')
        self.disponible_var.set(False)

        if not producto:
            messagebox.showwarning("Mensaje", "El producto no existe.")
            return

        self.nombre_var.set(producto.nombre)
        self.precio_var.set(producto.precio)
        self.cantidad_var.set(producto.cantidad)
        self.disponible_var.set(producto.disponible)

class ProductoCambiarDisponibilidadFrame(tk.Toplevel):
    def __init__(self, master, inventario):
        super().__init__(master)
        self.master = master

        self.inventario = inventario

        self.inicializar_gui()

    def inicializar_gui(self):
        self.title("Producto - Cambiar Disponibilidad")
        self.minsize(300, 200)
        
        label_codigo = ttk.Label(self, text="Código:")
        self.entry_codigo = ttk.Entry(self, validate="key", validatecommand=(self.register(self.validate_integer), "%P"))

        self.estado = tk.IntVar(value=0)
        self.checkbox_disponible = ttk.Checkbutton(self, text="¿Disponible para venta?", variable=self.estado)
        self.checkbox_disponible.state(['disabled'])

        # Evento para responder al cambio del checkbox:
        self.checkbox_disponible.bind("<Button-1>", self.cambiar_estado_producto)

        button_buscar = ttk.Button(self, text="Buscar", command=self.buscar_producto)

        label_codigo.grid(row=0, column=0, sticky=tk.W)
        self.entry_codigo.grid(row=0, column=1)
        button_buscar.grid(row=1, columnspan=2)
        self.checkbox_disponible.grid(row=2, columnspan=2, sticky=tk.W)

    def validate_integer(self, value):
        # Validar que solo se ingresen números enteros en el campo de código
        if value.isdigit() or value == "":
            return True
        return False

    def buscar_producto(self):
        codigo = self.entry_codigo.get().strip()

        if len(codigo) == 0:
            messagebox.showwarning("Mensaje", "El código es obligatorio.")
            return
        
        codigo = int(codigo)

        if not codigo and codigo == 0:
            messagebox.showwarning("Mensaje", "El código es obligatorio.")
            return
        
        producto = self.inventario.buscar_producto(codigo)

        if not producto:
            self.checkbox_disponible.state(['!disabled'])
            messagebox.showwarning("Mensaje", "El producto no existe.")
            return

        current_state = self.estado.get()
        new_state = 0 if current_state == 1 else 1
        self.estado.set(new_state)

        self.checkbox_disponible.state(['!disabled'])

    def cambiar_estado_producto(self, event):
        codigo = int(self.entry_codigo.get())

        if not codigo and codigo == 0:
            messagebox.showwarning("Mensaje", "El código es obligatorio.")
            return

        producto = self.inventario.buscar_producto(codigo)

        if not producto:
            messagebox.showwarning("Mensaje", "El producto no existe.")
            return

        self.inventario.cambiar_estado_producto(producto)

        self.inventario.guardar_datos()

class ReporteVentasRangoFechasFrame(tk.Toplevel):
    def __init__(self, master, inventario):
        super().__init__(master)
        self.inventario = inventario
        self.inicializar_gui()

    def inicializar_gui(self):
        hace_mes = datetime.datetime.now() - datetime.timedelta(days=30)
        hoy = datetime.datetime.now()

        patron_fecha = 'dd/mm/yyyy'
        
        tk.Label(self, text="Fecha de inicio (dd/mm/aaaa):").grid(row=0, column=0, padx=5, pady=5)
        self.dat_fecha_inicio = DateEntry(self, month=hace_mes.month, year=hace_mes.year, day=hace_mes.day, maxdate=hoy, date_pattern=patron_fecha)
        self.dat_fecha_inicio.grid(row=0, column=1, padx=5, pady=5)

        tk.Label(self, text="Fecha final (dd/mm/aaaa):").grid(row=1, column=0, padx=5, pady=5)
        self.dat_fecha_final = DateEntry(self, maxdate=hoy, date_pattern=patron_fecha)
        self.dat_fecha_final.grid(row=1, column=1, padx=5, pady=5)

        self.search_button = tk.Button(self, text="Buscar", command=self.buscar_ventas)
        self.search_button.grid(row=2, column=0, columnspan=2, padx=5, pady=5)

        self.tbl_ventas = tk.ttk.Treeview(self, columns=("id", 'nombre_producto', "fecha", "cantidad", "total"))
        self.tbl_ventas.grid(row=3, column=0, columnspan=2, padx=5, pady=5)

        self.tbl_ventas.heading("id", text="ID Producto")
        self.tbl_ventas.heading("nombre_producto", text="Nombre Producto")
        self.tbl_ventas.heading("fecha", text="Fecha")
        self.tbl_ventas.heading("cantidad", text="Cantidad")
        self.tbl_ventas.heading("total", text="Total")

    def buscar_ventas(self):
        fecha_inicio = self.dat_fecha_inicio.get()
        fecha_final = self.dat_fecha_final.get()

        if not self.es_fecha_valida(fecha_inicio):
            messagebox.showwarning("Advertencia", "El campo de Fecha de inicio debe tener el formato dd/mm/aaaa.")
            return
        
        if not self.es_fecha_valida(fecha_final):
            messagebox.showwarning("Advertencia", "El campo de Fecha final debe tener el formato dd/mm/aaaa.")
            return
        
        fecha_inicio = datetime.datetime.strptime(fecha_inicio, '%d/%m/%Y')
        fecha_final = datetime.datetime.strptime(fecha_final, '%d/%m/%Y')
        fecha_final = fecha_final.replace(hour=23, minute=59, second=59, microsecond=999999)

        ventas = self.inventario.ventas_rango_fecha(fecha_inicio, fecha_final)

        if len(ventas) == 0:
            messagebox.showinfo("Mensaje", "No se encontraron ventas en el rango de fechas especificado.")
            self.tbl_ventas.delete(*self.tbl_ventas.get_children())
            return

        for v in ventas:
            producto = self.inventario.buscar_producto(v.codigo_producto)
            fecha = v.fecha.strftime("%d/%B/%Y %H:%M:%S")

            total = "${:,.2f}".format(v.total_sin_iva * 1.19)

            self.tbl_ventas.insert("", tk.END, values=(v.codigo_producto, producto.nombre, fecha, v.cantidad, total))

    @staticmethod
    def es_fecha_valida(fecha_cadena):
        try:
            datetime.datetime.strptime(fecha_cadena, "%d/%m/%Y")
            return True
        except ValueError:
            return False

class Top5VendidosFrame(tk.Toplevel):
    def __init__(self, master, inventario, productos):
        super().__init__(master)
        self.inventario = inventario
        self.productos = productos

        self.inicializar_gui()
    
    def inicializar_gui(self):
        self.title("Tabla de Inventario")
        
        # Crear tabla
        self.tabla = ttk.Treeview(self)
        self.tabla['columns'] = ('codigo', 'nombre', 'precio', 'cantidad', 'total')
        
        # Configurar encabezados de columna
        self.tabla.heading('codigo', text='Código Producto')
        self.tabla.heading('nombre', text='Nombre Producto')
        self.tabla.heading('precio', text='Precio')
        self.tabla.heading('cantidad', text='Cantidad')
        self.tabla.heading('total', text='Total')
        
        # Configurar anchos de columna
        self.tabla.column('codigo', width=100)
        self.tabla.column('nombre', width=150)
        self.tabla.column('precio', width=80)
        self.tabla.column('cantidad', width=80)
        self.tabla.column('total', width=80)

        self.mostrar_tabla()

    def mostrar_tabla(self):

        for p in self.productos:
            producto = self.inventario.buscar_producto(p[0])
            cantidad = p[1]
            total = cantidad * producto.precio
            
            codigo = producto.codigo
            nombre = producto.nombre
            precio = producto.precio
            cantidad = producto.cantidad

            precio = "${:,.2f}".format(precio)
            total = "${:,.2f}".format(total)

            self.tabla.insert('', 'end', values=(codigo, nombre, precio, cantidad, total))
        
        self.tabla.pack(fill='both', expand=True)


producto = None

def main(page: Page):
    inventario = Inventario()
    page.title = "Gestor Inventario - Dispositivos s.a.s. - Por: OrtizOL"

    def close_dlg(e):
        dlg_modal.open = False
        page.update()
    
    def on_confirmar_cierre(e):
        page.window_close()

    dlg_modal = ft.AlertDialog(
        modal=True,
        title=ft.Text('Mensaje'),
        content=ft.Text(''),
        actions_alignment=ft.MainAxisAlignment.END
    )

    def mostrar_mensaje(mensaje):
        """
        Muestra un mensaje de advertencia en un diálogo modal.

        Parameters:
        mensaje: mensaje a mostrar en el diálogo modal.
        """
        dlg_modal.content = ft.Text(mensaje)
        dlg_modal.actions = [
            ft.TextButton("Aceptar", on_click=close_dlg),
        ]
        page.dialog = dlg_modal
        dlg_modal.open = True
        page.update()

    def on_click_salir(e):
        dlg_modal.content = ft.Text("¿Desea salir de la aplicación?")
        dlg_modal.actions = [
            ft.TextButton("Sí", on_click=on_confirmar_cierre),
            ft.TextButton("No", on_click=close_dlg),
        ]
        page.dialog = dlg_modal
        dlg_modal.open = True
        page.update()

    def on_click_nav_producto_registrar(e):
        page.go('/producto/registrar')
    
    def on_click_nav_producto_vender(e):
        page.go('/producto/vender')

    def on_click_nav_producto_buscar(e):
        page.go('/producto/buscar')

    def on_click_nav_producto_cambiar_disponibilidad(e):
        page.go('/producto/cambiar_disponibilidad')

    def on_click_nav_ventas_rango_fechas(e):
        page.go('/ventas/rango_fechas')

    def on_click_nav_producto_top_5_mas_vendidos(e):
        page.go('/producto/top_5_mas_vendidos')

    def on_click_nav_producto_top_5_menos_vendidos(e):
        page.go('/producto/top_5_menos_vendidos')

    def on_click_registrar_producto(e):
        codigo = txt_codigo.current.value.strip()
        nombre = txt_nombre.current.value.strip()
        precio = txt_precio.current.value.strip()
        cantidad = txt_cantidad.current.value.strip()
        disponible = chk_disponible_venta.current.value

        print('ID de thread evento:', threading.get_ident())

        if len(codigo) == 0 and len(nombre) == 0 and len(precio) == 0 and len(cantidad) == 0:
            mostrar_mensaje("Todos los campos son obligatorios. Los valores numéricos deben ser mayores a 0.")
            return
        
        try:
            codigo = int(codigo)
        except ValueError:
            mostrar_mensaje("El código debe ser numérico.")
            return
        
        if codigo <= 0:
            mostrar_mensaje("El código debe ser positivo.")
            return

        try:
            precio = float(precio)
        except ValueError:
            mostrar_mensaje("El precio debe ser numérico.")
            return

        if precio <= 0:
            mostrar_mensaje("El precio debe ser positivo.")
            return

        try:
            cantidad = int(cantidad)
        except ValueError:
            mostrar_mensaje("La cantidad debe ser numérica.")
            return

        if cantidad <= 0:
            mostrar_mensaje("La cantidad debe ser positiva.")
            return
        
        conexion = conectar('inventario/inventario.db')
        inventario.recibir_conexion_bd(conexion)
        producto = inventario.buscar_producto_por_codigo(codigo)

        if producto:
            mostrar_mensaje("Ya existe un producto con el código especificado.")
            return

        nuevo_producto = Producto(codigo, nombre, precio, cantidad, disponible)

        inventario.registrar_producto(nuevo_producto)

        mostrar_mensaje("El producto se ha creado de forma satisfactoria.")

        txt_codigo.current.value = ''
        txt_nombre.current.value = ''
        txt_precio.current.value = ''
        txt_cantidad.current.value = ''
        chk_disponible_venta.current.value = True

        conexion.close()

    txt_codigo = ft.Ref[ft.TextField]()
    txt_nombre = ft.Ref[ft.TextField]()
    txt_precio = ft.Ref[ft.TextField]()
    txt_cantidad = ft.Ref[ft.TextField]()
    chk_disponible_venta = ft.Ref[ft.Checkbox]()
    ref_txt_fecha_inicio = ft.Ref[ft.TextField]()
    ref_txt_fecha_fin = ft.Ref[ft.TextField]()
    ref_tbl_ventas = ft.Ref[ft.DataTable]()

    def on_change_cambiar_disponibilidad_producto(e):
        estado_disponibilidad = chk_disponible_venta.current.value

        producto.disponible = estado_disponibilidad

        conexion = conectar('inventario/inventario.db')
        inventario.recibir_conexion_bd(conexion)

        inventario.cambiar_estado_producto(producto.codigo, estado_disponibilidad)

        conexion.commit()

        conexion.close()

        mostrar_mensaje("El estado del producto se ha cambiado de forma satisfactoria.")

        page.update()

    def on_click_vender_producto(e):
        codigo = txt_codigo.current.value.strip()
        cantidad = txt_cantidad.current.value.strip()

        if len(codigo) == 0 or len(cantidad) == 0:
            mostrar_mensaje("Todos los campos son obligatorios. Los valores numéricos deben ser mayores a 0.")
            return
        
        try:
            codigo = int(codigo)
        except ValueError:
            mostrar_mensaje("El código debe ser numérico.")
            return
        
        if codigo <= 0:
            mostrar_mensaje("El código debe ser positivo.")
            return
        
        try:
            cantidad = int(cantidad)
        except ValueError:
            mostrar_mensaje("La cantidad debe ser numérica.")
            return

        if cantidad <= 0:
            mostrar_mensaje("La cantidad debe ser positiva.")
            return
        
        conexion = conectar('inventario/inventario.db')
        inventario.recibir_conexion_bd(conexion)

        producto = inventario.buscar_producto_por_codigo(codigo)

        if not producto:
            mostrar_mensaje("El producto no existe. Intente con otro código")
            return
        
        if not producto.disponible:
            mostrar_mensaje("El producto no está disponible para la venta.")
            return
        
        if cantidad > producto.cantidad:
            mostrar_mensaje("La cantidad solicitada es mayor a la cantidad disponible.")
            return
        
        total_sin_iva = producto.precio * cantidad

        venta = Venta(codigo, cantidad, total_sin_iva)

        conexion = conectar('inventario/inventario.db')
        inventario.recibir_conexion_bd(conexion)

        inventario.realizar_venta(venta)

        mostrar_mensaje("La venta se ha realizado de forma satisfactoria.")

        txt_codigo.current.value = ''
        txt_cantidad.current.value = ''
        
    def on_click_buscar_producto(e):
        codigo = txt_codigo.current.value.strip()

        if len(codigo) == 0:
            mostrar_mensaje("Todos los campos son obligatorios. Los valores numéricos deben ser mayores a 0.")
            return
        
        try:
            codigo = int(codigo)
        except ValueError:
            mostrar_mensaje("El código debe ser numérico.")
            return
        
        if codigo <= 0:
            mostrar_mensaje("El código debe ser positivo.")
            return
        
        conexion = conectar('inventario/inventario.db')
        inventario.recibir_conexion_bd(conexion)

        producto = inventario.buscar_producto_por_codigo(codigo)

        conexion.close()

        if not producto:
            mostrar_mensaje("El producto no existe.")
            txt_nombre.current.value = ''
            txt_precio.current.value = ''
            txt_cantidad.current.value = ''
            return

        txt_nombre.current.value = producto.nombre
        txt_precio.current.value = producto.precio
        txt_cantidad.current.value = producto.cantidad
        chk_disponible_venta.current.value = producto.disponible

        page.update()

    def on_click_buscar_producto_cambiar_disponibilidad(e):
        global producto
        
        codigo = txt_codigo.current.value.strip()

        if len(codigo) == 0:
            mostrar_mensaje("Todos los campos son obligatorios. Los valores numéricos deben ser mayores a 0.")
            return
        
        try:
            codigo = int(codigo)
        except ValueError:
            mostrar_mensaje("El código debe ser numérico.")
            return
        
        if codigo <= 0:
            mostrar_mensaje("El código debe ser positivo.")
            return
        
        conexion = conectar('inventario/inventario.db')
        inventario.recibir_conexion_bd(conexion)

        producto = inventario.buscar_producto_por_codigo(codigo)

        conexion.close()

        if not producto:
            mostrar_mensaje("El producto no existe.")
            chk_disponible_venta.current.value = False
            chk_disponible_venta.current.disabled = True
            return
        
        chk_disponible_venta.current.disabled = False
        chk_disponible_venta.current.value = bool(producto.disponible)
        chk_disponible_venta.current.update()

        page.update()

    def on_click_generar_reporte_ventas_rango_fechas(e):
        fecha_inicio = ref_txt_fecha_inicio.current.value.strip()
        fecha_fin = ref_txt_fecha_fin.current.value.strip()

        if len(fecha_inicio) == 0 or len(fecha_fin) == 0:
            mostrar_mensaje("Todos los campos son obligatorios.")
            return
        
        try:
            fecha_inicio = datetime.datetime.strptime(fecha_inicio, '%d/%m/%Y')
        except ValueError:
            mostrar_mensaje("El formato de la fecha de inicio es incorrecto. Debe ser dd/mm/aaaa.")
            return
        
        try:
            fecha_fin = datetime.datetime.strptime(fecha_fin, '%d/%m/%Y')
        except ValueError:
            mostrar_mensaje("El formato de la fecha final es incorrecto. Debe ser dd/mm/aaaa.")
            return
        
        if fecha_fin < fecha_inicio:
            mostrar_mensaje("La fecha final debe ser mayor o igual a la fecha de inicio.")
            return

        conexion = conectar('inventario/inventario.db')
        inventario.recibir_conexion_bd(conexion)

        ventas = inventario.ventas_rango_fechas(fecha_inicio, fecha_fin)

        conexion.close()

        if len(ventas) == 0:
            mostrar_mensaje("No se encontraron ventas en el rango de fechas especificado.")
            ref_tbl_ventas.current.rows = []
            ref_tbl_ventas.current.update()
            return

        rows = []

        for venta in ventas:
            cells = []
            cells.append(ft.DataCell(ft.Text(venta.codigo_producto)))
            cells.append(ft.DataCell(ft.Text(venta.fecha)))
            cells.append(ft.DataCell(ft.Text(venta.cantidad)))
            cells.append(ft.DataCell(ft.Text(venta.total_sin_iva * 1.19)))

            rows.append(ft.DataRow(cells=cells))

        ref_tbl_ventas.current.rows = rows
            
        ref_tbl_ventas.current.update()

    def generar_tabla():
        columnas = ('Código Producto', 'Fecha', 'Cantidad', 'Total')

        # Cree una lista de lista con 5 filas y 4 columnas:
        datos = [
            ['', '', '', ''],
        ]

        rows = []

        for fila in datos:
            cells = []

            for columna in fila:
                cells.append(ft.DataCell(ft.Text(columna)))

            rows.append(ft.DataRow(cells=cells))

        return ft.DataTable(
            columns=[
                ft.DataColumn(ft.Text(d)) for d in columnas
            ],
            rows=rows,
        )

    def generar_vista_producto_registrar():

        row_codigo = ft.ResponsiveRow([
            ft.Container(
                ft.TextField(
                    label="Código",
                    hint_text="Ingrese el código del producto",
                    ref=txt_codigo
                ),
                col={"sm": 12, "md": 12, "xl": 12},
            )
            ],
        )
        
        row_nombre = ft.ResponsiveRow([
            ft.Container(
                ft.TextField(
                    label="Nombre",
                    hint_text="Ingrese el nombre del producto",
                    ref=txt_nombre
                ),
                col={"sm": 12, "md": 12, "xl": 12},
            )
            ],
        )

        row_precio = ft.ResponsiveRow([
            ft.Container(
                ft.TextField(
                    label="Precio",
                    hint_text="Ingrese el precio del producto",
                    ref=txt_precio
                ),
                col={"sm": 12, "md": 12, "xl": 12},
            )
            ],
        )

        row_cantidad = ft.ResponsiveRow([
            ft.Container(
                ft.TextField(
                    label="Cantidad",
                    hint_text="Ingrese la cantidad del producto",
                    ref=txt_cantidad
                ),
                col={"sm": 12, "md": 12, "xl": 12},
            )
            ],
        )

        row_disponible_venta = ft.ResponsiveRow([
            ft.Container(
                ft.Checkbox(label="¿Disponible para venta?", value=True, ref=chk_disponible_venta),
                col={"sm": 12, "md": 12, "xl": 12},
            )
            ],
        )

        row_crear = ft.ResponsiveRow([
            ft.Container(
                ft.FilledButton(text='Crear', on_click=on_click_registrar_producto),
                col={"sm": 12, "md": 12, "xl": 12},
            )
            ],
        )

        return ft.Column(
            [
                row_codigo,
                row_nombre,
                row_precio,
                row_cantidad,
                row_disponible_venta,
                row_crear
            ],
            spacing=2
        )

    def generar_vista_producto_vender():
        row_codigo = ft.ResponsiveRow([
            ft.Container(
                ft.TextField(
                    label="Código",
                    hint_text="Ingrese el código del producto",
                    ref=txt_codigo
                ),
                col={"sm": 12, "md": 12, "xl": 12},
            )
            ],
        )

        row_cantidad = ft.ResponsiveRow([
            ft.Container(
                ft.TextField(
                    label="Cantidad",
                    hint_text="Ingrese la cantidad del producto",
                    ref=txt_cantidad
                ),
                col={"sm": 12, "md": 12, "xl": 12},
            )
            ],
        )

        row_vender = ft.ResponsiveRow([
            ft.Container(
                ft.FilledButton(text='Vender', on_click=on_click_vender_producto),
                col={"sm": 12, "md": 12, "xl": 12},
            )
            ],
        )

        return ft.Column(
            [
                row_codigo,
                row_cantidad,
                row_vender
            ],
            spacing=2
        )
        
    def generar_vista_producto_buscar():
        row_codigo = ft.ResponsiveRow([
            ft.Container(
                ft.TextField(
                    label="Código",
                    hint_text="Ingrese el código del producto a buscar",
                    ref=txt_codigo
                ),
                col={"sm": 12, "md": 12, "xl": 12},
            )
            ],
        )
        
        row_nombre = ft.ResponsiveRow([
            ft.Container(
                ft.TextField(
                    label="Nombre",
                    ref=txt_nombre,
                    read_only=True
                ),
                col={"sm": 12, "md": 12, "xl": 12},
            )
            ],
        )

        row_precio = ft.ResponsiveRow([
            ft.Container(
                ft.TextField(
                    label="Precio",
                    ref=txt_precio,
                    read_only=True
                ),
                col={"sm": 12, "md": 12, "xl": 12},
            )
            ],
        )

        row_cantidad = ft.ResponsiveRow([
            ft.Container(
                ft.TextField(
                    label="Cantidad",
                    ref=txt_cantidad,
                    read_only=True
                ),
                col={"sm": 12, "md": 12, "xl": 12},
            )
            ],
        )

        row_disponible_venta = ft.ResponsiveRow([
            ft.Container(
                ft.Checkbox(label="¿Disponible para venta?", value=True, ref=chk_disponible_venta),
                col={"sm": 12, "md": 12, "xl": 12},
            )
            ],
        )

        row_buscar = ft.ResponsiveRow([
            ft.Container(
                ft.FilledButton(text='Buscar', on_click=on_click_buscar_producto),
                col={"sm": 12, "md": 12, "xl": 12},
            )
            ],
        )

        return ft.Column(
            [
                row_codigo,
                row_buscar,
                row_nombre,
                row_precio,
                row_cantidad,
                row_disponible_venta
            ],
            spacing=2
        )

    def generar_vista_producto_cambiar_disponibilidad():
        row_codigo = ft.ResponsiveRow([
            ft.Container(
                ft.TextField(
                    label="Código",
                    hint_text="Ingrese el código del producto a buscar",
                    ref=txt_codigo
                ),
                col={"sm": 12, "md": 12, "xl": 12},
            )
            ],
        )

        row_buscar = ft.ResponsiveRow([
            ft.Container(
                ft.FilledButton(text='Buscar', on_click=on_click_buscar_producto_cambiar_disponibilidad),
                col={"sm": 12, "md": 12, "xl": 12},
            )
            ],
        )

        row_disponible_venta = ft.ResponsiveRow([
            ft.Container(
                ft.Checkbox(
                    label="¿Disponible para venta?", 
                    value=False, 
                    ref=chk_disponible_venta,
                    disabled=True,
                    on_change=on_change_cambiar_disponibilidad_producto
                    ),
                col={"sm": 12, "md": 12, "xl": 12},
            )
            ],
        )

        return ft.Column(
            [
                row_codigo,
                row_buscar,
                row_disponible_venta
            ],
            spacing=2
        )

    def generar_vista_reporte_ventas_rango_fechas():
        row_fecha_inicio = ft.ResponsiveRow([
            ft.Container(
                ft.TextField(
                    label="Fecha inicio",
                    hint_text="Ingrese la fecha de inicio",
                    ref=ref_txt_fecha_inicio
                ),
                col={"sm": 12, "md": 12, "xl": 12},
            )
            ],
        )
        
        row_fecha_fin = ft.ResponsiveRow([
            ft.Container(
                ft.TextField(
                    label="Fecha fin",
                    hint_text="Ingrese la fecha fin",
                    ref=ref_txt_fecha_fin
                ),
                col={"sm": 12, "md": 12, "xl": 12},
            )
            ],
        )

        row_buscar = ft.ResponsiveRow([
            ft.Container(
                ft.FilledButton(text='Buscar', on_click=on_click_generar_reporte_ventas_rango_fechas),
                col={"sm": 12, "md": 12, "xl": 12},
            )
            ],
        )

        ref_tbl_ventas.current = generar_tabla()

        row_ventas = ft.ResponsiveRow([
            ft.Container(
                ref_tbl_ventas.current,
                col={"sm": 12, "md": 12, "xl": 12},
            )
            ],
        )

        return ft.Column(
            [
                row_fecha_inicio,
                row_fecha_fin,
                row_buscar,
                row_ventas
            ],
            spacing=2
        )

    def generar_vista_reporte_top_5(orden='descendete'):
        rows = []

        conexion = conectar('inventario/inventario.db')
        inventario.recibir_conexion_bd(conexion)

        if orden == 'descendete':
            productos = inventario.top_5_mas_vendidos()
        else:
            productos = inventario.top_5_menos_vendidos()

        for p in productos:
            producto = inventario.buscar_producto_por_codigo(p[0])

            cells = []

            cells.append(ft.DataCell(ft.Text(producto.codigo)))
            cells.append(ft.DataCell(ft.Text(producto.nombre)))
            cells.append(ft.DataCell(ft.Text(producto.precio)))
            cells.append(ft.DataCell(ft.Text(p[1])))

            total = producto.precio * p[1] * 1.19
            cells.append(ft.DataCell(ft.Text(total)))

            rows.append(ft.DataRow(cells=cells))

        conexion.close()
        
        ref_tbl_ventas.current = ft.DataTable(
            columns=[
                ft.DataColumn(ft.Text("Código Producto")),
                ft.DataColumn(ft.Text("Nombre Producto")),
                ft.DataColumn(ft.Text("Precio")),
                ft.DataColumn(ft.Text("Cantidad vendida")),
                ft.DataColumn(ft.Text("Total")),
            ],
            rows=rows,
        )

        row_ventas = ft.ResponsiveRow([
            ft.Container(
                ref_tbl_ventas.current,
                col={"sm": 12, "md": 12, "xl": 12},
            )
            ],
        )

        return ft.Column(
            [
                row_ventas
            ],
            spacing=2
        )

    def on_route_change(e):
        page.views.clear()

        mnu_archivo = PopupMenuButton(
            content=Text('Archivo'),
            items=[
                ft.PopupMenuItem(text='Salir', on_click=on_click_salir)
            ]
        )

        mnu_productos = PopupMenuButton(
            content=Text('Productos'),
            items=[
                ft.PopupMenuItem(text='Registrar', on_click=on_click_nav_producto_registrar),
                ft.PopupMenuItem(text='Vender', on_click=on_click_nav_producto_vender),
                ft.PopupMenuItem(text='Buscar', on_click=on_click_nav_producto_buscar),
                ft.PopupMenuItem(text='Cambiar disponibilidad', on_click=on_click_nav_producto_cambiar_disponibilidad)
            ]
        )
        
        mnu_reportes = PopupMenuButton(
            content=Text('Reportes'),
            items=[
                ft.PopupMenuItem(text='Ventas en un rango de fechas', on_click=on_click_nav_ventas_rango_fechas),
                ft.PopupMenuItem(text='Top 5 productos más vendidos', on_click=on_click_nav_producto_top_5_mas_vendidos),
                ft.PopupMenuItem(text='Top 5 productos menos vendidos', on_click=on_click_nav_producto_top_5_menos_vendidos),
            ]
        )

        mbr_principal = ft.ResponsiveRow([
            ft.Container(
                mnu_archivo,
                col={"sm": 2, "md": 2, "xl": 2},
            ),
            ft.Container(
                mnu_productos,
                col={"sm": 2, "md": 2, "xl": 2},
            ),
            ft.Container(
                mnu_reportes,
                col={"sm": 2, "md": 2, "xl": 2},
            )
            ],
        )

        page.views.append(
            View(
                "/",
                [
                    AppBar(title=Text('Gestión Inventario - Dispositivos S.a.s.')),
                    mbr_principal
                ],
            )
        )

        if page.route == "/producto/registrar":
            page.views.append(
                View(
                    "/producto/registrar",
                    [
                        AppBar(title=Text("Producto - Registrar"), bgcolor=colors.SURFACE_VARIANT),
                        generar_vista_producto_registrar()
                    ]
                )
            )
        
        if page.route == "/producto/vender":
            page.views.append(
                View(
                    "/producto/vender",
                    [
                        AppBar(title=Text("Producto - Vender"), bgcolor=colors.SURFACE_VARIANT),
                        generar_vista_producto_vender()
                    ]
                )
            )
        
        if page.route == "/producto/buscar":
            page.views.append(
                View(
                    "/producto/buscar",
                    [
                        AppBar(title=Text("Producto - Buscar"), bgcolor=colors.SURFACE_VARIANT),
                        generar_vista_producto_buscar()
                    ]
                )
            )
        
        if page.route == "/producto/cambiar_disponibilidad":
            page.views.append(
                View(
                    "/producto/cambiar_disponibilidad",
                    [
                        AppBar(title=Text("Producto - Cambiar Disponibilidad"), bgcolor=colors.SURFACE_VARIANT),
                        generar_vista_producto_cambiar_disponibilidad()
                    ]
                )
            )
        
        if page.route == '/ventas/rango_fechas':
            page.views.append(
                View(
                    '/ventas/rango_fechas',
                    [
                        AppBar(title=Text("Reporte - Ventas en un rango de fechas"), bgcolor=colors.SURFACE_VARIANT),
                        generar_vista_reporte_ventas_rango_fechas()
                    ]
                )
            )
        
        if page.route == '/producto/top_5_mas_vendidos':
            page.views.append(
                View(
                    '/producto/top_5_mas_vendidos',
                    [
                        AppBar(title=Text("Reporte - Top 5 Productos Más Vendidos"), bgcolor=colors.SURFACE_VARIANT),
                        generar_vista_reporte_top_5()
                    ]
                )
            )
        
        if page.route == '/producto/top_5_menos_vendidos':
            page.views.append(
                View(
                    '/producto/top_5_menos_vendidos',
                    [
                        AppBar(title=Text("Reporte - Top 5 Productos Menos Vendidos"), bgcolor=colors.SURFACE_VARIANT),
                        generar_vista_reporte_top_5('ascendente')
                    ]
                )
            )

        page.update()

    def view_pop(e):
        print("View pop:", e.view)
        page.views.pop()
        top_view = page.views[-1]
        page.go(top_view.route)

    page.on_route_change = on_route_change
    page.on_view_pop = view_pop

    page.go(page.route)


if __name__ == "__main__":
    ft.app(target=main)
